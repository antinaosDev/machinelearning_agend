import streamlit as st
import pandas as pd
import numpy as np
import joblib
import io
import csv
from datetime import datetime
import base64

st.set_page_config(page_title="A2S Praedix - Predicción Inteligente", page_icon="image.png", layout="wide")

# --- CSS PERSONALIZADO ---
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
    * { font-family: 'Inter', sans-serif; }
    :root {
        --surface: #f7f9fb; --primary: #022448; --primary-container: #1e3a5f;
        --secondary: #0051d5; --error: #ba1a1a; --error-container: #ffdad6;
        --tertiary: #002252; --on-surface: #191c1e; --on-surface-variant: #43474e;
        --surface-container-lowest: #ffffff; --surface-container-low: #f2f4f6;
        --surface-container-high: #e6e8ea; --outline-variant: #c4c6cf;
    }
    .main { background-color: var(--surface); }
    .glass-nav { background: rgba(255,255,255,0.8); backdrop-filter: blur(12px); border-bottom: 1px solid var(--outline-variant); }
    .card-shadow { box-shadow: 0 4px 24px -2px rgba(2,36,72,0.06); }
    .main-cta-gradient { background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); }
    .status-card { background: var(--surface-container-lowest); padding: 1.5rem; border-radius: 1rem; border: 1px solid var(--outline-variant); }
    .total-card { background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); padding: 1.5rem; border-radius: 1rem; color: white; }
    .result-card { background: var(--surface-container-lowest); border: 2px solid var(--primary-container); border-radius: 1rem; padding: 2rem; }
    .story-card { background: var(--surface-container-lowest); padding: 2rem; border-radius: 1rem; border: 1px solid var(--outline-variant); }
    .insight-card { background: var(--surface-container-low); padding: 1.5rem; border-radius: 1rem; }
    .priority-badge { padding: 0.25rem 0.75rem; border-radius: 9999px; font-size: 0.625rem; font-weight: 900; text-transform: uppercase; }
    .priority-critical { background: #ffdad6; color: #93000a; }
    .priority-high { background: rgba(0,81,213,0.1); color: #0051d5; }
    .priority-medium { background: #e6e8ea; color: #43474e; }
    .priority-low { background: #dcfce7; color: #166534; }
    .hero-gradient { background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); padding: 4rem; border-radius: 1rem; color: white; }
    .metric-card { background: white; border-radius: 1rem; padding: 1.5rem; border: 1px solid #e6e8ea; }
    .metric-value { font-size: 2.5rem; font-weight: 900; color: #022448; }
    .metric-label { font-size: 0.75rem; font-weight: 700; color: #43474e; text-transform: uppercase; letter-spacing: 0.05em; }
    .progress-bar { background: #e6e8ea; border-radius: 0.5rem; height: 12px; overflow: hidden; }
    .progress-fill { height: 100%; border-radius: 0.5rem; transition: width 0.5s ease; }
    .info-box { background: #eff6ff; border-left: 4px solid #0051d5; padding: 1rem; border-radius: 0 0.5rem 0.5rem 0; }
    .warning-box { background: #fff7ed; border-left: 4px solid #f59e0b; padding: 1rem; border-radius: 0 0.5rem 0.5rem 0; }
    .success-box { background: #f0fdf4; border-left: 4px solid #16a534; padding: 1rem; border-radius: 0 0.5rem 0.5rem 0; }
    .download-card { background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); border-radius: 1rem; padding: 1.5rem; color: white; cursor: pointer; transition: transform 0.2s; }
    .download-card:hover { transform: scale(1.02); }
    .upload-zone { border: 2px dashed #c4c6cf; background: #fafbfc; border-radius: 1rem; padding: 3rem; text-align: center; transition: all 0.3s ease; cursor: pointer; }
    .upload-zone:hover { border-color: #0051d5; background: #eff6ff; }
    .upload-zone.dragover { border-color: #0051d5; background: #dbeafe; }
    </style>
    """, unsafe_allow_html=True)

# --- CARGA DEL MODELO ---
@st.cache_resource
def cargar_modelo():
    try:
        return joblib.load('modelo_rf_agenda_cholchol.pkl')
    except:
        return None

modelo = cargar_modelo()

columnas_ia = [
    'GENERO_MASCULINO', 'PROCEDENCIA_RURAL', 'ES DISCAPACITADA_SI',
    'ES SENAME_SI', 'ES EMBARAZADA_SI', 'ESCOLARIDAD_SIN INSTRUCCION',
    'ESCOLARIDAD_PREBASICA', 'ESCOLARIDAD_BASICA', 'ESCOLARIDAD_ESPECIAL O DIFERENCIAL',
    'ESCOLARIDAD_MEDIA', 'ESCOLARIDAD_TECNICO DE NIVEL SUPERIOR',
    'ESCOLARIDAD_EDUCACION SUPERIOR', 'PREVISION_FONASA - A',
    'PREVISION_FONASA - B', 'PREVISION_FONASA - C', 'PREVISION_FONASA - D',
    'MES_AGEND_Enero', 'MES_AGEND_Febrero', 'MES_AGEND_Marzo',
    'MES_AGEND_Abril', 'MES_AGEND_Mayo', 'MES_AGEND_Junio',
    'MES_AGEND_Julio', 'MES_AGEND_Agosto', 'MES_AGEND_Septiembre',
    'MES_AGEND_Octubre', 'MES_AGEND_Noviembre', 'MES_AGEND_Diciembre',
    'BLOQUE_HORARIO_Mañana', 'BLOQUE_HORARIO_Mediodía', 'BLOQUE_HORARIO_Tarde',
    'DIA_AGEND_Lunes', 'DIA_AGEND_Martes', 'DIA_AGEND_Miércoles',
    'DIA_AGEND_Jueves', 'DIA_AGEND_Viernes', 'DIA_AGEND_Sábado',
    'DIA_AGEND_Domingo'
]

# --- FUNCIONES DE TRANSFORMACIÓN ---
def limpiar_datos(df):
    for col in df.select_dtypes(include=['object', 'string']).columns:
        df[col] = df[col].astype(str).str.upper().str.strip()
    df = df.drop_duplicates()
    df = df.replace(['NaN', 'nan', ' ', 'None', 'NAN'], np.nan)
    return df

def formatear_hora(hora_str):
    if pd.isna(hora_str) or hora_str == 'NAN': return np.nan
    try:
        return int(str(hora_str).split(':')[0])
    except:
        return np.nan

def clasificar_bloque_horario(hora_num):
    if pd.isna(hora_num): return 'Sin Dato'
    if 0 <= hora_num < 8: return 'Nocturno/Madrugada'
    elif 8 <= hora_num < 13: return 'Mañana'
    elif 13 <= hora_num < 15: return 'Mediodía'
    elif 15 <= hora_num < 17: return 'Tarde'
    elif 17 <= hora_num < 20: return 'Extensión Horaria'
    elif 20 <= hora_num <= 23: return 'Noche'
    return 'Sin Dato'

def obtener_mes_espanol(fecha):
    try:
        meses = {1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril', 5:'Mayo', 6:'Junio',
                 7:'Julio', 8:'Agosto', 9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'}
        return meses.get(fecha.month, 'Sin Dato')
    except:
        return 'Sin Dato'

def obtener_dia_espanol(fecha):
    try:
        dias = {'Monday':'Lunes', 'Tuesday':'Martes', 'Wednesday':'Miércoles', 
                'Thursday':'Jueves', 'Friday':'Viernes', 'Saturday':'Sábado', 'Sunday':'Domingo'}
        return dias.get(fecha.day_name(), 'Sin Dato')
    except:
        return 'Sin Dato'

def estandarizar_categorias(df):
    mapeos = {
        'GENERO': {'MUJER': 'FEMENINO', 'HOMBRE': 'MASCULINO'},
        'ESCOLARIDAD': {'MEDIA O SECUNDARIA': 'MEDIA', 'BASICA O PRIMARIA': 'BASICA', 'PROFESIONAL UNIVERSITARIO': 'EDUCACION SUPERIOR'},
        'PROCEDENCIA': {'-': 'SIN INFORMACION', 'NAN': 'SIN INFORMACION'}
    }
    for col, reemplazo in mapeos.items():
        if col in df.columns:
            df[col] = df[col].replace(reemplazo)
    return df

def procesar_dataframe(df):
    df = limpiar_datos(df)
    df_work = df.copy()
    df_work.columns = df_work.columns.str.upper().str.strip()
    
    mapeo_cols = {'SEXO': 'GENERO', 'ZONA': 'PROCEDENCIA', 'LUGAR': 'PROCEDENCIA',
                  'FECHA ASIGNADA': 'FECHA_AGENDADA', 'HORA ASIGNADA': 'HORA_AGENDADA', 'FECHA': 'FECHA_AGENDADA', 'HORA': 'HORA_AGENDADA'}
    for viejo, nuevo in mapeo_cols.items():
        if viejo in df_work.columns and nuevo not in df_work.columns:
            df_work[nuevo] = df_work[viejo]
    
    col_fecha = next((c for c in df_work.columns if 'FECHA' in c and 'AGEND' in c), None)
    if col_fecha:
        df_work['FECHA_PARSED'] = pd.to_datetime(df_work[col_fecha], errors='coerce', dayfirst=True)
        df_work['MES_CALCULADO'] = df_work['FECHA_PARSED'].apply(lambda x: obtener_mes_espanol(x) if pd.notna(x) else 'Sin Dato')
        df_work['DIA_CALCULADO'] = df_work['FECHA_PARSED'].apply(lambda x: obtener_dia_espanol(x) if pd.notna(x) else 'Sin Dato')
    else:
        df_work['MES_CALCULADO'] = 'Sin Dato'
        df_work['DIA_CALCULADO'] = 'Sin Dato'
    
    col_hora = next((c for c in df_work.columns if 'HORA' in c and 'AGEND' in c), None)
    if col_hora:
        df_work['HORA_NUM'] = df_work[col_hora].apply(formatear_hora)
    else:
        df_work['HORA_NUM'] = np.nan
    
    df_work['BLOQUE_CALCULADO'] = df_work['HORA_NUM'].apply(clasificar_bloque_horario)
    df_work = estandarizar_categorias(df_work)
    
    cols_necesarias = ['GENERO', 'PROCEDENCIA', 'PREVISION', 'ESCOLARIDAD', 'ES DISCAPACITADA', 'ES SENAME', 'ES EMBARAZADA']
    for col in cols_necesarias:
        if col not in df_work.columns:
            df_work[col] = 'SIN INFORMACION'
    
    if 'PROCEDENCIA' not in df_work.columns or df_work['PROCEDENCIA'].isna().all():
        df_work['PROCEDENCIA'] = 'RURAL'
    if 'PREVISION' not in df_work.columns or df_work['PREVISION'].isna().all():
        df_work['PREVISION'] = 'FONASA - A'
    if 'GENERO' not in df_work.columns or df_work['GENERO'].isna().all():
        df_work['GENERO'] = 'MASCULINO'
    if 'ESCOLARIDAD' not in df_work.columns or df_work['ESCOLARIDAD'].isna().all():
        df_work['ESCOLARIDAD'] = 'BASICA'
    
    df_work['PROCEDENCIA'] = df_work['PROCEDENCIA'].fillna('RURAL').replace(['NAN', 'SIN INFORMACION'], 'RURAL')
    df_work['PREVISION'] = df_work['PREVISION'].fillna('FONASA - A').replace(['NAN', 'SIN INFORMACION'], 'FONASA - A')
    df_work['GENERO'] = df_work['GENERO'].fillna('MASCULINO').replace(['NAN'], 'MASCULINO')
    df_work['ESCOLARIDAD'] = df_work['ESCOLARIDAD'].fillna('BASICA').replace(['NAN'], 'BASICA')
    
    for col in cols_necesarias:
        if col not in df_work.columns:
            df_work[col] = 'NO'
    
    df['MES_CALCULADO'] = df_work['MES_CALCULADO']
    df['DIA_CALCULADO'] = df_work['DIA_CALCULADO']
    df['BLOQUE_CALCULADO'] = df_work['BLOQUE_CALCULADO']
    df['PROCEDENCIA'] = df_work['PROCEDENCIA']
    df['GENERO'] = df_work['GENERO']
    df['PREVISION'] = df_work['PREVISION']
    df['ESCOLARIDAD'] = df_work['ESCOLARIDAD']
    
    return df

def obtener_prioridad(prob):
    if prob >= 70: return "CRÍTICO", "Llamada obligatoria inmediata"
    if prob >= 50: return "ALTO", "Confirmar vía telefónica"
    if prob >= 25: return "MEDIO", "Enviar recordatorio WhatsApp"
    return "BAJO", "Monitoreo estándar"

def obtener_clase_prioridad(prob):
    if prob >= 70: return "priority-critical"
    if prob >= 50: return "priority-high"
    if prob >= 25: return "priority-medium"
    return "priority-low"

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crear_plantilla_excel():
    """Crea una plantilla de Excel con validaciones, formato institucional y estilos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    
    # Colores institucionales
    COLOR_PRIMARIO = "022448"  # Azul oscuro
    COLOR_FONDO = "f7f9fb"  # Gris claro
    
    # Fuentes sin bordes
    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_datos = Font(size=10)
    fill_header = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    fill_ejemplo = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    fill_vacio = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    columnas = [
        'RUT', 'NOMBRE', 'PATERNO', 'MATERNO', 'DOMICILIO', 'GENERO', 'PROCEDENCIA', 
        'PREVISION', 'ESCOLARIDAD', 'FECHA ASIGNADA', 'HORA ASIGNADA',
        'ES DISCAPACITADA', 'ES SENAME', 'ES EMBARAZADA', 'TELEFONO1'
    ]
    
    # Escribir encabezados con formato
    for col, header in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar filas con formato (hasta fila 1000)
    for row in range(2, 1002):
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            if row == 2:  # Fila de ejemplo
                cell.fill = fill_ejemplo
            else:
                cell.fill = fill_vacio
            cell.font = font_datos
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir validaciones de datos (dropdowns) - OBLIGATORIAS
    dv_genero = DataValidation(type="list", formula1='"MASCULINO,FEMENINO"', allow_blank=False)
    dv_genero.error = 'Debe seleccionar un valor de la lista'
    dv_genero.errorTitle = 'Campo Obligatorio'
    dv_genero.prompt = 'Seleccione una opción'
    dv_genero.promptTitle = 'Género'
    ws.add_data_validation(dv_genero)
    dv_genero.add('F2:F1001')
    
    dv_procedencia = DataValidation(type="list", formula1='"RURAL,URBANO"', allow_blank=False)
    dv_procedencia.prompt = 'Seleccione una opción'
    dv_procedencia.promptTitle = 'Procedencia'
    ws.add_data_validation(dv_procedencia)
    dv_procedencia.add('G2:G1001')
    
    dv_prevision = DataValidation(type="list", formula1='"FONASA - A,FONASA - B,FONASA - C,FONASA - D"', allow_blank=False)
    dv_prevision.prompt = 'Seleccione una opción'
    dv_prevision.promptTitle = 'Previsión'
    ws.add_data_validation(dv_prevision)
    dv_prevision.add('H2:H1001')
    
    dv_escolaridad = DataValidation(type="list", formula1='"BASICA,MEDIA,SUPERIOR,SIN INSTRUCCION,PREBASICA,TECNICO DE NIVEL SUPERIOR,ESPECIAL O DIFERENCIAL,NO RESPONDE,NO RECUERDA"', allow_blank=False)
    dv_escolaridad.prompt = 'Seleccione una opción'
    dv_escolaridad.promptTitle = 'Escolaridad'
    ws.add_data_validation(dv_escolaridad)
    dv_escolaridad.add('I2:I1001')
    
    # Validación de fecha (formato DD/MM/AAAA)
    dv_fecha = DataValidation(type="custom", formula1='AND(ISNUMBER(SEARCH("/",J2)),LEN(J2)=10)', allow_blank=False)
    dv_fecha.error = 'Formato debe ser DD/MM/AAAA (ej: 15/05/2024)'
    dv_fecha.errorTitle = 'Formato de Fecha Inválido'
    ws.add_data_validation(dv_fecha)
    dv_fecha.add('J2:J1001')
    
    # Validación de hora (formato HH:MM)
    dv_hora = DataValidation(type="custom", formula1='AND(ISNUMBER(SEARCH(":",K2)),LEN(K2)=5)', allow_blank=False)
    dv_hora.error = 'Formato debe ser HH:MM (ej: 08:30)'
    dv_hora.errorTitle = 'Formato de Hora Inválido'
    ws.add_data_validation(dv_hora)
    dv_hora.add('K2:K1001')
    
    dv_si_no = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    dv_si_no.prompt = 'Seleccione SI o NO'
    dv_si_no.promptTitle = 'Confirmación'
    ws.add_data_validation(dv_si_no)
    dv_si_no.add('L2:L1001')
    dv_si_no.add('M2:M1001')
    dv_si_no.add('N2:N1001')
    
    # Ajustar ancho de columnas
    anchos = [15, 20, 15, 15, 25, 12, 12, 12, 28, 15, 10, 18, 12, 15, 15]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    
    # Fijar primera fila
    ws.freeze_panes = 'A2'
    
    # Quitar líneas de cuadrícula
    ws.sheet_view.showGridLines = False
    
    return wb

# --- NAVEGACIÓN PROFESIONAL ---
if 'current_screen' not in st.session_state:
    st.session_state.current_screen = 'CargaMasiva'

st.markdown("""
    <style>
    .nav-container {
        background: white;
        border-radius: 1rem;
        padding: 0.75rem;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        margin-bottom: 1rem;
    }
    .nav-logo-container {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        padding: 0.5rem 1rem;
        border-bottom: 1px solid #e6e8ea;
        margin-bottom: 0.5rem;
    }
    .nav-logo-img {
        width: 40px;
        height: 40px;
    }
    .nav-logo-text {
        font-size: 1.25rem;
        font-weight: 900;
        color: #022448;
    }
    .nav-logo-sub {
        font-size: 0.7rem;
        color: #0051d5;
        font-weight: 500;
    }
    .nav-tabs {
        display: flex;
        gap: 0.25rem;
    }
    .nav-tab {
        padding: 0.75rem 1.25rem;
        border-radius: 0.5rem;
        font-weight: 600;
        font-size: 0.875rem;
        cursor: pointer;
        transition: all 0.2s ease;
        text-align: center;
        flex: 1;
    }
    .nav-tab:hover {
        background: #f2f4f6;
    }
    .nav-tab.active {
        background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%);
        color: white;
    }
    .nav-tab.inactive {
        color: #43474e;
    }
    </style>
""", unsafe_allow_html=True)

# Mostrar navegación con tabs visuales
st.markdown('<div class="nav-container">', unsafe_allow_html=True)

logo_col1, logo_col2 = st.columns([1, 8])
with logo_col1:
    st.image("image.png", width=80)
with logo_col2:
    st.markdown("""
        <div style="padding-top: 0.5rem;">
            <div style="font-size: 1.25rem; font-weight: 900; color: #022448;">A2S Praedix</div>
            <div style="font-size: 0.7rem; color: #0051d5; font-weight: 500;">Predicción inteligente de asistencia médica</div>
        </div>
    """, unsafe_allow_html=True)

nav_tabs = [
    ("📊 Carga Masiva", "CargaMasiva"),
    ("👤 Registro Individual", "RegistroIndividual"),
    ("📖 Historia del Proyecto", "HistoriaProyecto"),
    ("📋 Plantilla & Guía", "PlantillaGuia")
]

tabs_col = st.columns(4)
for i, (label, screen_id) in enumerate(nav_tabs):
    with tabs_col[i]:
        if st.button(label, key=f"nav_{screen_id}", use_container_width=True, 
                    type="primary" if st.session_state.current_screen == screen_id else "secondary"):
            st.session_state.current_screen = screen_id
            st.rerun()

st.markdown('</div>', unsafe_allow_html=True)
st.markdown("<div style='height: 1rem;'></div>", unsafe_allow_html=True)

# ==================== PANTALLA 1: CARGA MASIVA ====================
if st.session_state.current_screen == 'CargaMasiva':
    st.markdown("""
        <div style="margin-bottom: 2rem;">
            <h1 style="font-size: 2.25rem; font-weight: 900; color: #191c1e;">Carga Masiva</h1>
            <p style="color: #43474e; font-size: 1.125rem;">Gestión centralizada de registros clínicos y análisis de riesgo poblacional.</p>
        </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown("""<div class="status-card" style="border-left: 4px solid #ba1a1a;"><div style="display: flex; justify-content: space-between; align-items: start;"><div><p class="metric-label">Riesgo Crítico (≥80%)</p><h2 style="font-size: 2.25rem; font-weight: 900; color: #191c1e;">91.1%</h2></div><div style="padding: 0.5rem; border-radius: 0.75rem; background: #ffdad6;"><span style="color: #ba1a1a; font-size: 1.25rem;">🔴</span></div></div><p style="margin-top: 1rem; font-size: 0.7rem; color: #43474e; opacity: 0.8;">Paciente rural citado 08:30 AM</p></div>""", unsafe_allow_html=True)
    with col2:
        st.markdown("""<div class="status-card" style="border-left: 4px solid #0051d5;"><div style="display: flex; justify-content: space-between; align-items: start;"><div><p class="metric-label">Riesgo Alto (60-80%)</p><h2 style="font-size: 2.25rem; font-weight: 900; color: #191c1e;">~60%</h2></div><div style="padding: 0.5rem; border-radius: 0.75rem; background: rgba(0,81,213,0.1);"><span style="color: #0051d5; font-size: 1.25rem;">🟠</span></div></div><p style="margin-top: 1rem; font-size: 0.7rem; color: #43474e; opacity: 0.8;">Hombre rural, horario laboral</p></div>""", unsafe_allow_html=True)
    with col3:
        st.markdown("""<div class="status-card" style="border-left: 4px solid #002252;"><div style="display: flex; justify-content: space-between; align-items: start;"><div><p class="metric-label">Inasistencia General</p><h2 style="font-size: 2.25rem; font-weight: 900; color: #191c1e;">14.9%</h2></div><div style="padding: 0.5rem; border-radius: 0.75rem; background: rgba(0,34,82,0.1);"><span style="color: #002252; font-size: 1.25rem;">📊</span></div></div><p style="margin-top: 1rem; font-size: 0.7rem; color: #43474e; opacity: 0.8;">Promedio CESFAM Cholchol</p></div>""", unsafe_allow_html=True)
    with col4:
        st.markdown("""<div class="total-card"><div style="display: flex; justify-content: space-between; align-items: start;"><div><p style="font-size: 0.75rem; color: #93c5fd;">Registros Históricos</p><h2 style="font-size: 2.25rem; font-weight: 700;">430,183</h2></div><div style="padding: 0.5rem; border-radius: 0.5rem; background: rgba(255,255,255,0.2);"><span style="color: white; font-size: 1.25rem;">🗄️</span></div></div><p style="margin-top: 1rem; font-size: 0.7rem; color: #93c5fd;">Horas Médicas 2016-2026 | Precisión: 85.45%</p></div>""", unsafe_allow_html=True)
    
    st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
    
    # === SECCIÓN DE CARGA DE ARCHIVO MEJORADA ===
    st.markdown("""
        <div style="background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); padding: 2rem; border-radius: 1rem; margin-bottom: 1.5rem;">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <div>
                    <h3 style="font-size: 1.25rem; font-weight: 700; color: white; margin-bottom: 0.5rem;">📁 Subir Archivos</h3>
                    <p style="font-size: 0.875rem; color: #93c5fd;">Carga tu archivo de agenda para analizar riesgo de inasistencia</p>
                </div>
                <div style="display: flex; gap: 0.75rem;">
                    <span style="background: rgba(255,255,255,0.2); padding: 0.5rem 1rem; border-radius: 0.5rem; color: white; font-size: 0.8rem;">📊 .xlsx</span>
                    <span style="background: rgba(255,255,255,0.2); padding: 0.5rem 1rem; border-radius: 0.5rem; color: white; font-size: 0.8rem;">📄 .csv</span>
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)
    
    # CREAR PLANTILLA DESCARGABLE CON DROPDOWNS
    wb_plantilla = crear_plantilla_excel()
    buffer_plantilla = io.BytesIO()
    wb_plantilla.save(buffer_plantilla)
    buffer_plantilla.seek(0)
    plantilla_data = buffer_plantilla.getvalue()
    
    # MOSTRAR BOTÓN DE DESCARGA DE PLANTILLA
    st.markdown("""
        <div style="background: #f0fdf4; border: 1px solid #16a534; border-radius: 0.75rem; padding: 1rem; margin-bottom: 1.5rem; display: flex; justify-content: space-between; align-items: center;">
            <div style="display: flex; align-items: center; gap: 1rem;">
                <div style="width: 2.5rem; height: 2.5rem; border-radius: 0.5rem; background: #16a534; display: flex; align-items: center; justify-content: center;">📥</div>
                <div>
                    <p style="font-weight: 600; color: #166534; margin: 0;">¿No tienes un archivo?</p>
                    <p style="font-size: 0.8rem; color: #43474e; margin: 0;">Descarga la plantilla de ejemplo</p>
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)
    
    st.download_button(
        label="📥 Descargar Plantilla Ejemplo (.xlsx)",
        data=plantilla_data,
        file_name="Plantilla_Rescate_Cholchol.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Descarga una plantilla con el formato correcto"
    )
    
    st.markdown("<div style='height: 1rem;'></div>", unsafe_allow_html=True)
    
    # ZONA DE CARGA VISUAL
    if modelo is None:
        st.error("Archivo de modelo (.pkl) no encontrado."); st.stop()
    
    archivo = st.file_uploader(
        "Arrastra tu archivo de agenda aquí o haz clic para seleccionar",
        type=["xlsx", "csv", "xls"],
        label_visibility="collapsed"
    )
    
    # VISUALIZACIÓN DE ESTADO DE CARGA
    if archivo is not None:
        st.markdown(f"""
            <div style="background: #eff6ff; border: 1px solid #0051d5; border-radius: 1rem; padding: 1.5rem; margin-top: 1rem;">
                <div style="display: flex; align-items: center; gap: 1rem;">
                    <div style="width: 3rem; height: 3rem; border-radius: 50%; background: #0051d5; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.5rem;">✓</div>
                    <div>
                        <p style="font-weight: 700; color: #022448; margin: 0;">Archivo cargado exitosamente</p>
                        <p style="font-size: 0.875rem; color: #43474e; margin: 0;">📄 {archivo.name} | 📊 {archivo.size/1024:.1f} KB</p>
                    </div>
                </div>
            </div>
        """, unsafe_allow_html=True)
    
    if archivo:
        try:
            if archivo.name.endswith('.csv'):
                df_orig = pd.read_csv(archivo, encoding='latin-1', sep=None, engine='python', on_bad_lines='skip')
            else:
                df_orig = pd.read_excel(archivo)
            
            if not df_orig.empty:
                with st.spinner("🧠 IA Analizando datos..."):
                    df_orig = procesar_dataframe(df_orig)
                    
                    cols_ia = ['GENERO', 'PROCEDENCIA', 'PREVISION', 'ESCOLARIDAD', 
                               'MES_CALCULADO', 'DIA_CALCULADO', 'BLOQUE_CALCULADO']
                    presentes = [c for c in cols_ia if c in df_orig.columns]
                    
                    df_work = df_orig[presentes].copy()
                    for c in presentes:
                        df_work[c] = df_work[c].astype(str).str.upper().str.strip().replace('NAN', 'SIN DATO')
                    
                    X_input = pd.get_dummies(df_work).reindex(columns=columnas_ia, fill_value=0)
                    
                    probs = modelo.predict_proba(X_input)[:, 1]
                    df_orig['RIESGO_%'] = (probs * 100).round(1)
                    res = [obtener_prioridad(p) for p in df_orig['RIESGO_%']]
                    df_orig['PRIORIDAD'] = [r[0] for r in res]
                    df_orig['PLAN_ACCION'] = [r[1] for r in res]
                    df_orig['CLASE_PRIORIDAD'] = [obtener_clase_prioridad(p) for p in df_orig['RIESGO_%']]
                    
                    df_orig = df_orig.sort_values('RIESGO_%', ascending=False)
                
                st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
                
                st.markdown("""<div style="background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); padding: 2rem; border-radius: 1rem; margin-bottom: 2rem; color: white;"><h3 style="font-size: 1.25rem; font-weight: 700; margin-bottom: 0.5rem;">📈 Resultados del Análisis - Archivo Procesado</h3><p style="font-size: 0.875rem; color: #93c5fd;">Métricas calculadas en tiempo real</p></div>""", unsafe_allow_html=True)
                
                total_analizadas = len(df_orig)
                criticos = len(df_orig[df_orig['RIESGO_%'] >= 70])
                altos = len(df_orig[(df_orig['RIESGO_%'] >= 50) & (df_orig['RIESGO_%'] < 70)])
                medios = len(df_orig[(df_orig['RIESGO_%'] >= 25) & (df_orig['RIESGO_%'] < 50)])
                bajos = len(df_orig[df_orig['RIESGO_%'] < 25])
                riesgo_promedio = df_orig['RIESGO_%'].mean()
                
                proc_counts = df_orig['PROCEDENCIA'].value_counts() if 'PROCEDENCIA' in df_orig.columns else pd.Series()
                rural_count = proc_counts.get('RURAL', 0)
                urbano_count = proc_counts.get('URBANO', 0)
                bloque_counts = df_orig['BLOQUE_CALCULADO'].value_counts() if 'BLOQUE_CALCULADO' in df_orig.columns else pd.Series()
                
                kpi1_col1, kpi1_col2, kpi1_col3, kpi1_col4 = st.columns(4)
                with kpi1_col1:
                    st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">Total Registros</p><p class="metric-value" style="color: #022448;">{total_analizadas:,}</p><p style="font-size: 0.75rem; color: #16a534;">100% Procesados</p></div>""", unsafe_allow_html=True)
                with kpi1_col2:
                    st.markdown(f"""<div class="metric-card" style="text-align: center; border-left: 4px solid #ba1a1a;"><p class="metric-label">🔴 Crítico (≥70%)</p><p class="metric-value" style="color: #ba1a1a;">{criticos:,}</p><p style="font-size: 0.75rem; color: #ba1a1a;">{round(criticos/total_analizadas*100,1)}% del total</p></div>""", unsafe_allow_html=True)
                with kpi1_col3:
                    st.markdown(f"""<div class="metric-card" style="text-align: center; border-left: 4px solid #0051d5;"><p class="metric-label">🟠 Alto (50-69%)</p><p class="metric-value" style="color: #0051d5;">{altos:,}</p><p style="font-size: 0.75rem; color: #0051d5;">{round(altos/total_analizadas*100,1)}% del total</p></div>""", unsafe_allow_html=True)
                with kpi1_col4:
                    st.markdown(f"""<div class="metric-card" style="text-align: center; border-left: 4px solid #002252;"><p class="metric-label">🟡 Medio (25-49%)</p><p class="metric-value" style="color: #002252;">{medios:,}</p><p style="font-size: 0.75rem; color: #002252;">{round(medios/total_analizadas*100,1)}% del total</p></div>""", unsafe_allow_html=True)
                
                st.markdown("<div style='height: 1rem;'></div>", unsafe_allow_html=True)
                
                kpi2_col1, kpi2_col2, kpi2_col3, kpi2_col4 = st.columns(4)
                with kpi2_col1:
                    st.markdown(f"""<div class="metric-card" style="text-align: center; border-left: 4px solid #16a534;"><p class="metric-label">🟢 Bajo (&lt;25%)</p><p class="metric-value" style="color: #16a534;">{bajos:,}</p><p style="font-size: 0.75rem; color: #16a534;">{round(bajos/total_analizadas*100,1)}% del total</p></div>""", unsafe_allow_html=True)
                with kpi2_col2:
                    st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">Riesgo Promedio</p><p class="metric-value" style="color: #022448;">{riesgo_promedio:.1f}%</p><p style="font-size: 0.75rem; color: #43474e;">del archivo cargado</p></div>""", unsafe_allow_html=True)
                with kpi2_col3:
                    st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">📍 Pacientes Rurales</p><p class="metric-value" style="color: #ba1a1a;">{rural_count:,}</p><p style="font-size: 0.75rem; color: #ba1a1a;">{round(rural_count/total_analizadas*100,1)}% del total</p></div>""", unsafe_allow_html=True)
                with kpi2_col4:
                    st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">🏙️ Pacientes Urbanos</p><p class="metric-value" style="color: #0051d5;">{urbano_count:,}</p><p style="font-size: 0.75rem; color: #0051d5;">{round(urbano_count/total_analizadas*100,1)}% del total</p></div>""", unsafe_allow_html=True)
                
                st.markdown("<div style='height: 1.5rem;'></div>", unsafe_allow_html=True)
                
                if not bloque_counts.empty:
                    st.markdown("""<h4 style="font-size: 1.125rem; font-weight: 700; color: #022448; margin-bottom: 1rem;">🕐 Distribución por Bloque Horario</h4>""", unsafe_allow_html=True)
                    
                    blo_col1, blo_col2 = st.columns([2, 1])
                    with blo_col1:
                        for bloque, count in bloque_counts.items():
                            pct = round(count/total_analizadas*100, 1)
                            color = "#ba1a1a" if bloque in ['Mañana'] else "#0051d5" if bloque in ['Tarde'] else "#16a534"
                            st.markdown(f"""<div style="margin-bottom: 0.75rem;"><div style="display: flex; justify-content: space-between; margin-bottom: 0.25rem;"><span style="font-weight: 600; color: #191c1e;">{bloque}</span><span style="font-weight: 700; color: {color};">{count:,} ({pct}%)</span></div><div style="background: #e6e8ea; border-radius: 0.5rem; height: 8px;"><div style="background: {color}; border-radius: 0.5rem; height: 100%; width: {pct}%;"></div></div></div>""", unsafe_allow_html=True)
                    with blo_col2:
                        max_bloque = bloque_counts.idxmax() if not bloque_counts.empty else "N/A"
                        st.markdown(f"""<div style="background: #f2f4f6; padding: 1rem; border-radius: 0.75rem;"><h5 style="font-size: 0.875rem; font-weight: 700; color: #022448; margin-bottom: 0.75rem;">📌 Insight</h5><p style="font-size: 0.8rem; color: #43474e;">Mayor concentración de citas en: <strong style="color: #022448;">{max_bloque}</strong></p></div>""", unsafe_allow_html=True)
                
                st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
                
                st.markdown("### 📋 Lista de Pacientes por Prioridad", unsafe_allow_html=True)
                cols_v = ['PRIORIDAD', 'RIESGO_%', 'DIA_CALCULADO', 'BLOQUE_CALCULADO', 'PROCEDENCIA']
                cols_exist = [c for c in cols_v if c in df_orig.columns]
                if cols_exist:
                    st.dataframe(df_orig[cols_exist].head(50), use_container_width=True, hide_index=True)
                
                # Columnas necesarias para lista de rescate/llamado
                # Primero preservar todas las columnas originales del archivo
                df_rescate = df_orig.copy()
                
                # Agregar recomendaciones basadas en el riesgo
                def obtener_recomendacion(row):
                    riesgo = row.get('RIESGO_%', 0)
                    if riesgo >= 70:
                        return "LLAMADA OBLIGATORIA - Contactar inmediatamente para confirmar asistencia"
                    elif riesgo >= 50:
                        return "Confirmar vía telefónica o enviar WhatsApp"
                    elif riesgo >= 25:
                        return "Enviar recordatorio automático"
                    else:
                        return "Monitoreo estándar"
                
                df_rescate['RECOMENDACIÓN'] = df_rescate.apply(obtener_recomendacion, axis=1)
                
                # Asegurar que existan las columnas necesarias (solo las esenciales)
                cols_originales = ['RUT', 'NOMBRE', 'PATERNO', 'MATERNO', 'GENERO', 'TELEFONO1',
                                  'DIRECCION', 'PRIORIDAD', 'RIESGO_%', 'FECHA_ASIGNADA', 'HORA_ASIGNADA', 
                                  'PROCEDENCIA', 'MES_CALCULADO', 'DIA_CALCULADO', 'BLOQUE_CALCULADO', 'PREVISION',
                                  'RECOMENDACIÓN']
                
                for col in cols_originales:
                    if col not in df_rescate.columns:
                        if col == 'PRIORIDAD':
                            df_rescate[col] = df_rescate.apply(lambda x: obtener_prioridad(x.get('RIESGO_%', 0))[0] if 'RIESGO_%' in x else 'BAJO', axis=1)
                        elif col == 'RIESGO_%':
                            df_rescate[col] = 0.0
                        else:
                            df_rescate[col] = ''
                
                # Agregar columna de notas vacía para preenchimiento
                df_rescate['NOTAS / ACCIÓN REALIZADA'] = ''
                
                # Seleccionar columnas精简izadas para el export
                cols_export = ['RUT', 'NOMBRE', 'PATERNO', 'MATERNO', 'GENERO', 'TELEFONO1',
                              'DIRECCION', 'PRIORIDAD', 'RIESGO_%', 'FECHA_ASIGNADA', 'HORA_ASIGNADA',
                              'PROCEDENCIA', 'MES_CALCULADO', 'DIA_CALCULADO', 'BLOQUE_CALCULADO', 'PREVISION', 
                              'RECOMENDACIÓN', 'NOTAS / ACCIÓN REALIZADA']
                
                cols_finales = [c for c in cols_export if c in df_rescate.columns]
                df_rescate = df_rescate[cols_finales]
                
                # Ordenar por riesgo (mayor primero)
                if 'RIESGO_%' in df_rescate.columns:
                    df_rescate = df_rescate.sort_values('RIESGO_%', ascending=False)
                
                # Crear Excel con formato institucional
                wb_rescate = Workbook()
                ws = wb_rescate.active
                ws.title = "Lista Rescate"
                
                COLOR_PRIMARIO = "022448"
                font_header = Font(bold=True, color="FFFFFF", size=11)
                font_datos = Font(size=10)
                fill_header = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
                fill_alto = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                fill_medio = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                fill_bajo = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                
                # Encabezados
                for col, header in enumerate(df_rescate.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Datos con colores por riesgo y prioridad
                for row_idx, row in enumerate(df_rescate.values, 2):
                    riesgo = df_rescate.iloc[row_idx-2].get('RIESGO_%', 0) if 'RIESGO_%' in df_rescate.columns else 0
                    prioridad = df_rescate.iloc[row_idx-2].get('PRIORIDAD', '').upper() if 'PRIORIDAD' in df_rescate.columns else ''
                    
                    for col_idx, valor in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                        cell.font = font_datos
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Color por prioridad
                        if prioridad == 'CRÍTICO' or riesgo >= 70:
                            cell.fill = fill_alto
                        elif prioridad == 'ALTO' or riesgo >= 50:
                            cell.fill = fill_medio
                        else:
                            cell.fill = fill_bajo
                
                # Ancho columnas - ajustado para las columnas simplificadas
                anchos = [15, 20, 15, 15, 12, 15, 30, 15, 10, 15, 10, 15, 12, 12, 15, 12, 45, 30]
                for col, ancho in enumerate(anchos[:len(df_rescate.columns)], 1):
                    ws.column_dimensions[get_column_letter(col)].width = ancho
                
                ws.freeze_panes = 'A2'
                ws.sheet_view.showGridLines = False
                
                buffer = io.BytesIO()
                wb_rescate.save(buffer)
                buffer.seek(0)
                st.download_button(label="📥 Descargar Lista de Rescate", data=buffer.getvalue(), file_name="Lista_Rescate_Cholchol.xlsx", type="primary")
                
        except Exception as e:
            st.error(f"Error de procesamiento: {e}")
    
    st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
    bento_col1, bento_col2 = st.columns([2, 1])
    with bento_col1:
        st.markdown("""<div style="background: white; padding: 1.5rem; border-radius: 1rem; display: flex; gap: 1.5rem; border: 1px solid #c4c6cf;"><div style="width: 33%; border-radius: 0.75rem; overflow: hidden; flex-shrink: 0;"><img src="https://images.unsplash.com/photo-1448375240586-882707db888b?w=600&q=80" style="width: 100%; height: 100%; object-fit: cover;" referrerPolicy="no-referrer" /></div><div style="display: flex; flex-direction: column; justify-content: center; gap: 1rem;"><h4 style="font-size: 1.25rem; font-weight: 700; color: #022448;">Inteligencia Predictiva</h4><p style="color: #43474e; line-height: 1.6;">Random Forest con 85.45% de precisión. Analiza género, procedencia, previsión, escolaridad, mes, día y bloque horario.</p></div></div>""", unsafe_allow_html=True)
    with bento_col2:
        st.markdown("""<div style="background: linear-gradient(135deg, #1e3a5f 0%, #022448 100%); padding: 2rem; border-radius: 1rem; color: white; display: flex; flex-direction: column; justify-content: space-between; height: 100%;"><div style="font-size: 3rem;">🛡️</div><div><h4 style="font-size: 1.25rem; font-weight: 700;">Seguridad de Datos</h4><p style="font-size: 0.875rem; color: #93c5fd;">Los datos se procesan localmente. No se almacenan en la nube.</p></div></div>""", unsafe_allow_html=True)

# ==================== PANTALLA 2: REGISTRO INDIVIDUAL ====================
elif st.session_state.current_screen == 'RegistroIndividual':
    st.markdown("""
        <div style="margin-bottom: 2rem;">
            <h1 style="font-size: 2.25rem; font-weight: 900; color: #191c1e;">Registro Individual</h1>
            <p style="color: #43474e; font-size: 1.125rem;">Evaluación de riesgo de inasistencia por paciente</p>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="info-box" style="margin-bottom: 2rem;">
            <h4 style="font-weight: 700; color: #022448; margin-bottom: 0.5rem;">📋 Instrucciones de Uso</h4>
            <p style="color: #43474e; font-size: 0.9rem;">Complete los datos del paciente para obtener una predicción de riesgo de inasistencia. El sistema utiliza el modelo Random Forest entrenado con 430,183 registros históricos.</p>
        </div>
    """, unsafe_allow_html=True)
    
    if modelo is None:
        st.error("Archivo de modelo (.pkl) no encontrado."); st.stop()
    
    col_form1, col_form2, col_form3 = st.columns([1, 1, 1])
    with col_form1:
        st.markdown("""<div style="background: white; padding: 1.5rem; border-radius: 1rem; border: 1px solid #e6e8ea; margin-bottom: 1rem;"><div style="display: flex; align-items: center; gap: 0.75rem; margin-bottom: 1.5rem;"><div style="width: 2.5rem; height: 2.5rem; border-radius: 0.75rem; background: #022448; display: flex; align-items: center; justify-content: center; color: white;">👤</div><div><h3 style="font-size: 1rem; font-weight: 700; color: #022448;">Datos del Paciente</h3><p style="font-size: 0.75rem; color: #43474e;">Información sociodemográfica</p></div></div></div>""", unsafe_allow_html=True)
        
        g_i = st.selectbox("Género", ["MASCULINO", "FEMENINO"], help="Seleccione el género del paciente")
        p_i = st.selectbox("Procedencia", ["RURAL", "URBANO", "SIN INFORMACION"], help="Zona de residencia del paciente")
        pr_i = st.selectbox("Previsión", ["FONASA - A", "FONASA - B", "FONASA - C", "FONASA - D"], help="Sistema de salud")
        es_i = st.selectbox("Escolaridad", ["BASICA", "MEDIA", "SUPERIOR", "SIN INSTRUCCION", "PREBASICA", "TECNICO DE NIVEL SUPERIOR"], help="Nivel educacional")
    
    with col_form2:
        st.markdown("""<div style="background: white; padding: 1.5rem; border-radius: 1rem; border: 1px solid #e6e8ea; margin-bottom: 1rem;"><div style="display: flex; align-items: center; gap: 0.75rem; margin-bottom: 1.5rem;"><div style="width: 2.5rem; height: 2.5rem; border-radius: 0.75rem; background: #0051d5; display: flex; align-items: center; justify-content: center; color: white;">🕐</div><div><h3 style="font-size: 1rem; font-weight: 700; color: #022448;">Datos de la Hora Agendada</h3><p style="font-size: 0.75rem; color: #43474e;">Parámetros de agendamiento</p></div></div></div>""", unsafe_allow_html=True)
        
        me_i = st.selectbox("Mes", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"])
        di_i = st.selectbox("Día", ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"])
        bl_i = st.selectbox("Bloque", ["Mañana", "Mediodía", "Tarde", "Nocturno/Madrugada", "Extensión Horaria", "Noche"])
    
    with col_form3:
        st.markdown("""<div style="background: white; padding: 1.5rem; border-radius: 1rem; border: 1px solid #e6e8ea; margin-bottom: 1rem;"><div style="display: flex; align-items: center; gap: 0.75rem; margin-bottom: 1.5rem;"><div style="width: 2.5rem; height: 2.5rem; border-radius: 0.75rem; background: #002252; display: flex; align-items: center; justify-content: center; color: white;">⚠️</div><div><h3 style="font-size: 1rem; font-weight: 700; color: #022448;">Condiciones Especiales</h3><p style="font-size: 0.75rem; color: #43474e;">Factores de vulnerabilidad</p></div></div></div>""", unsafe_allow_html=True)
        
        dis_i = st.radio("¿Discapacidad?", ["NO", "SI"], horizontal=True)
        sen_i = st.radio("¿SENAME?", ["NO", "SI"], horizontal=True)
        emb_i = st.radio("¿Embarazo?", ["NO", "SI"], horizontal=True)
    
    st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
    
    btn = st.button("🎯 Calcular Probabilidad de Inasistencia", type="primary", use_container_width=True)
    
    if btn:
        df_i = pd.DataFrame({
            'GENERO': [g_i.upper()], 'PROCEDENCIA': [p_i.upper()], 'PREVISION': [pr_i.upper()], 
            'ESCOLARIDAD': [es_i.upper()], 'MES_CALCULADO': [me_i.capitalize()], 
            'DIA_CALCULADO': [di_i.capitalize()], 'BLOQUE_CALCULADO': [bl_i.capitalize()]
        })
        X_i = pd.get_dummies(df_i).reindex(columns=columnas_ia, fill_value=0)
        p_res = (modelo.predict_proba(X_i)[0][1] * 100).round(1)
        prio_res, plan_res = obtener_prioridad(p_res)
        clase_prio = obtener_clase_prioridad(p_res)
        
        st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
        
        st.markdown(f"""<div style="background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); padding: 2.5rem; border-radius: 1rem; color: white; margin-bottom: 2rem; text-align: center;"><h3 style="font-size: 1rem; font-weight: 600; color: #93c5fd; margin-bottom: 1rem;">PROBABILIDAD DE INASISTENCIA</h3><p style="font-size: 4rem; font-weight: 900; color: white;">{p_res}%</p><span style="background: {'#ffdad6' if p_res>=70 else 'rgba(0,81,213,0.2)' if p_res>=50 else '#e6e8ea' if p_res>=25 else '#dcfce7'}; padding: 0.5rem 1.5rem; border-radius: 2rem; font-weight: 700; color: {'#93000a' if p_res>=70 else '#0051d5' if p_res>=50 else '#43474e' if p_res>=25 else '#166534'};">{prio_res}</span></div>""", unsafe_allow_html=True)
        
        res_col1, res_col2 = st.columns([1, 1])
        with res_col1:
            st.markdown("""<div style="background: white; padding: 1.5rem; border-radius: 1rem; border: 1px solid #e6e8ea;"><h4 style="font-size: 1rem; font-weight: 700; color: #022448; margin-bottom: 1rem;">💡 Plan de Acción</h4>""", unsafe_allow_html=True)
            if p_res >= 70:
                st.markdown("""<div class="warning-box"><p style="font-weight: 600; color: #92400e;">🔴 Intervención Urgente</p><p style="font-size: 0.875rem; color: #43474e;">Llamada telefónica obligatoria inmediata.</p></div>""", unsafe_allow_html=True)
            elif p_res >= 50:
                st.markdown("""<div class="info-box"><p style="font-weight: 600; color: #022448;">🟠 Seguimiento Priorizado</p><p style="font-size: 0.875rem; color: #43474e;">Confirmar asistencia vía telefónica.</p></div>""", unsafe_allow_html=True)
            elif p_res >= 25:
                st.markdown("""<div class="success-box"><p style="font-weight: 600; color: #166534;">🟡 Recordatorio Preventivo</p><p style="font-size: 0.875rem; color: #43474e;">Enviar recordatorio por WhatsApp.</p></div>""", unsafe_allow_html=True)
            else:
                st.markdown("""<div class="success-box"><p style="font-weight: 600; color: #166534;">🟢 Monitoreo Estándar</p><p style="font-size: 0.875rem; color: #43474e;">Seguimiento regular.</p></div>""", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
        
        with res_col2:
            factores_html = ""
            for k in [f"Origen: {p_i}", f"Previsión: {pr_i}", f"Bloque: {bl_i}", f"Mes: {me_i}", f"Día: {di_i}"]:
                factores_html += f'<li style="display: flex; align-items: center; gap: 0.75rem; padding: 0.5rem 0; border-bottom: 1px solid #e6e8ea;"><span style="width: 8px; height: 8px; border-radius: 50%; background: #022448;"></span>{k}</li>'
            st.markdown(f"""<div style="background: white; padding: 1.5rem; border-radius: 1rem; border: 1px solid #e6e8ea;"><h4 style="font-size: 1rem; font-weight: 700; color: #022448; margin-bottom: 1rem;">📊 Factores de Riesgo</h4><ul style="list-style: none; padding: 0; margin: 0;">{factores_html}</ul></div>""", unsafe_allow_html=True)
        
        st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
        
        comp_col1, comp_col2, comp_col3, comp_col4 = st.columns(4)
        with comp_col1:
            st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">Vs. Promedio Cholchol</p><p style="font-size: 1.5rem; font-weight: 700; color: {'#ba1a1a' if p_res > 14.9 else '#16a534'}>{'+' if p_res > 14.9 else ''}{round(p_res - 14.9, 1)}%</p><p style="font-size: 0.7rem; color: #43474e;">vs 14.9% base</p></div>""", unsafe_allow_html=True)
        with comp_col2:
            st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">Mujeres (15.8%)</p><p style="font-size: 1.5rem; font-weight: 700; color: {'#ba1a1a' if p_res > 15.8 else '#16a534'}>{'+' if p_res > 15.8 else ''}{round(p_res - 15.8, 1)}%</p><p style="font-size: 0.7rem; color: #43474e;">vs género</p></div>""", unsafe_allow_html=True)
        with comp_col3:
            st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">Hombres (13.4%)</p><p style="font-size: 1.5rem; font-weight: 700; color: {'#ba1a1a' if p_res > 13.4 else '#16a534'}>{'+' if p_res > 13.4 else ''}{round(p_res - 13.4, 1)}%</p><p style="font-size: 0.7rem; color: #43474e;">vs género</p></div>""", unsafe_allow_html=True)
        with comp_col4:
            st.markdown(f"""<div class="metric-card" style="text-align: center;"><p class="metric-label">T. Edad (10.5%)</p><p style="font-size: 1.5rem; font-weight: 700; color: {'#ba1a1a' if p_res > 10.5 else '#16a534'}>{'+' if p_res > 10.5 else ''}{round(p_res - 10.5, 1)}%</p><p style="font-size: 0.7rem; color: #43474e;">vs grupo vulnerable</p></div>""", unsafe_allow_html=True)

# ==================== PANTALLA 3: HISTORIA DEL PROYECTO ====================
elif st.session_state.current_screen == 'HistoriaProyecto':
    st.markdown("""
        <div style="margin-bottom: 3rem;">
            <h1 style="font-size: 2.25rem; font-weight: 900; color: #191c1e;">Historia del Proyecto</h1>
            <p style="color: #43474e; font-size: 1.125rem;">Evolución y hallazgos del sistema de predicción de inasistencias CESFAM Cholchol</p>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""<div class="hero-gradient" style="margin-bottom: 3rem; position: relative; overflow: hidden;"><div style="position: relative; z-index: 10; max-width: 40rem;"><h1 style="font-size: 2.5rem; font-weight: 900; letter-spacing: -0.02em; line-height: 1.2;">Transformando la Gestión Clínica</h1><p style="font-size: 1rem; color: #93c5fd; margin-top: 1rem; line-height: 1.6;">De 430,183 registros dispersos a un sistema predictivo con 85.45% de precisión</p></div><div style="position: absolute; right: 2rem; top: 50%; transform: translateY(-50%); font-size: 8rem; opacity: 0.1;">📊</div></div>""", unsafe_allow_html=True)
    
    story_col1, story_col2, story_col3 = st.columns(3)
    with story_col1:
        st.markdown("""<div class="story-card" style="border-top: 4px solid #ba1a1a;"><div style="width: 3rem; height: 3rem; border-radius: 50%; background: #ffdad6; display: flex; align-items: center; justify-content: center; margin-bottom: 1rem;"><span style="font-size: 1.25rem;">⚠️</span></div><h4 style="font-size: 1.125rem; font-weight: 700; color: #191c1e; margin-bottom: 0.75rem;">Fase 1: Diagnóstico</h4><p style="color: #43474e; line-height: 1.5; font-size: 0.9rem;">14.9% de inasistencia. 430,183 registros sin capacidad predictiva.</p><div style="margin-top: 1rem; padding: 0.75rem; background: #fff7ed; border-radius: 0.5rem;"><p style="font-size: 0.75rem; font-weight: 600; color: #92400e;">Problema Central</p><p style="font-size: 0.8rem; color: #43474e;">12 fuentes de datos distintas</p></div></div>""", unsafe_allow_html=True)
    with story_col2:
        st.markdown("""<div class="story-card" style="border-top: 4px solid #0051d5;"><div style="width: 3rem; height: 3rem; border-radius: 50%; background: rgba(0,81,213,0.1); display: flex; align-items: center; justify-content: center; margin-bottom: 1rem;"><span style="font-size: 1.25rem;">🩺</span></div><h4 style="font-size: 1.125rem; font-weight: 700; color: #191c1e; margin-bottom: 0.75rem;">Fase 2: Desarrollo</h4><p style="color: #43474e; line-height: 1.5; font-size: 0.9rem;">ETL iterativo y Random Forest con class_weight='balanced'.</p><div style="margin-top: 1rem; padding: 0.75rem; background: #eff6ff; border-radius: 0.5rem;"><p style="font-size: 0.75rem; font-weight: 600; color: #1e40af;">Solución Técnica</p><p style="font-size: 0.8rem; color: #43474e;">36 features, validación cruzada</p></div></div>""", unsafe_allow_html=True)
    with story_col3:
        st.markdown("""<div class="story-card" style="border-top: 4px solid #16a534;"><div style="width: 3rem; height: 3rem; border-radius: 50%; background: #dcfce7; display: flex; align-items: center; justify-content: center; margin-bottom: 1rem;"><span style="font-size: 1.25rem;">📁</span></div><h4 style="font-size: 1.125rem; font-weight: 700; color: #191c1e; margin-bottom: 0.75rem;">Fase 3: Impacto</h4><p style="color: #43474e; line-height: 1.5; font-size: 0.9rem;">Lista de Rescate, 15% reducción tiempos administrativos.</p><div style="margin-top: 1rem; padding: 0.75rem; background: #f0fdf4; border-radius: 0.5rem;"><p style="font-size: 0.75rem; font-weight: 600; color: #166534;">Resultado</p><p style="font-size: 0.8rem; color: #43474e;">85.45% precisión</p></div></div>""", unsafe_allow_html=True)
    
    st.markdown("<div style='height: 3rem;'></div>", unsafe_allow_html=True)
    
    st.markdown("""<h3 style="font-size: 1.5rem; font-weight: 800; color: #022448; margin-bottom: 1.5rem;">📈 Resultados del Modelamiento</h3>""", unsafe_allow_html=True)
    
    col_model1, col_model2, col_model3, col_model4 = st.columns(4)
    with col_model1:
        st.markdown("""<div class="metric-card" style="text-align: center; border: 2px solid #16a534;"><p class="metric-label">Random Forest</p><p style="font-size: 2rem; font-weight: 900; color: #16a534;">85.45%</p><p style="font-size: 0.7rem; color: #43474e;">Mejor modelo</p></div>""", unsafe_allow_html=True)
    with col_model2:
        st.markdown("""<div class="metric-card" style="text-align: center;"><p class="metric-label">Gradient Boosting</p><p style="font-size: 2rem; font-weight: 900; color: #43474e;">85.34%</p><p style="font-size: 0.7rem; color: #43474e;">Rápido</p></div>""", unsafe_allow_html=True)
    with col_model3:
        st.markdown("""<div class="metric-card" style="text-align: center;"><p class="metric-label">Árbol Decisión</p><p style="font-size: 2rem; font-weight: 900; color: #43474e;">85.30%</p><p style="font-size: 0.7rem; color: #43474e;">Interpretable</p></div>""", unsafe_allow_html=True)
    with col_model4:
        st.markdown("""<div class="metric-card" style="text-align: center;"><p class="metric-label">Regresión Logística</p><p style="font-size: 2rem; font-weight: 900; color: #43474e;">85.03%</p><p style="font-size: 0.7rem; color: #43474e;">Base</p></div>""", unsafe_allow_html=True)
    
    st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
    
    st.markdown("""<h3 style="font-size: 1.5rem; font-weight: 800; color: #022448; margin-bottom: 1.5rem;">🎯 Factores de Predicción (Top 10)</h3>""", unsafe_allow_html=True)
    
    factores_data = [
        ("PROCEDENCIA_RURAL", "9.41%", "#ba1a1a"),
        ("GENERO_MASCULINO", "8.03%", "#ba1a1a"),
        ("BLOQUE_HORARIO_Mañana", "5.56%", "#f59e0b"),
        ("MES_AGEND_Diciembre", "5.44%", "#f59e0b"),
        ("BLOQUE_HORARIO_Tarde", "5.28%", "#f59e0b"),
        ("ESCOLARIDAD_MEDIA", "4.25%", "#0051d5"),
        ("ESCOLARIDAD_BASICA", "3.92%", "#0051d5"),
        ("PREVISION_FONASA - B", "3.84%", "#0051d5"),
        ("PREVISION_FONASA - A", "3.72%", "#0051d5"),
        ("DIA_AGEND_Viernes", "3.70%", "#0051d5"),
    ]
    
    for i, (factor, valor, color) in enumerate(factores_data):
        st.markdown(f"""<div style="display: flex; align-items: center; margin-bottom: 0.75rem; padding: 1rem; background: #f7f9fb; border-radius: 0.5rem; border-left: 4px solid {color};"><div style="flex: 1;"><p style="font-size: 0.9rem; font-weight: 700; color: #191c1e; margin: 0;">{factor}</p></div><div style="text-align: right;"><p style="font-size: 1.25rem; font-weight: 900; color: {color}; margin: 0;">{valor}</p></div></div>""", unsafe_allow_html=True)
    
    st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
    
    st.markdown("""<h3 style="font-size: 1.5rem; font-weight: 800; color: #022448; margin-bottom: 1.5rem;">📊 Tasas de Inasistencia por Categoría</h3>""", unsafe_allow_html=True)
    
    tab_tasa1, tab_tasa2, tab_tasa3, tab_tasa4 = st.tabs(["Horario", "Territorio", "Previsión", "Ciclo Vital"])
    
    with tab_tasa1:
        st.markdown("""<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 1rem;">
            <div style="padding: 1rem; background: #fff7ed; border-radius: 0.5rem; border-left: 4px solid #f59e0b;"><p style="font-weight: 700; margin: 0;">Mañana</p><p style="font-size: 1.5rem; font-weight: 900; margin: 0; color: #f59e0b;">16.3%</p></div>
            <div style="padding: 1rem; background: #fff7ed; border-radius: 0.5rem; border-left: 4px solid #f59e0b;"><p style="font-weight: 700; margin: 0;">Tarde</p><p style="font-size: 1.5rem; font-weight: 900; margin: 0; color: #f59e0b;">16.2%</p></div>
            <div style="padding: 1rem; background: #dcfce7; border-radius: 0.5rem; border-left: 4px solid #16a534;"><p style="font-weight: 700; margin: 0;">Mediodía</p><p style="font-size: 1.5rem; font-weight: 900; margin: 0; color: #16a534;">12.8%</p></div>
            <div style="padding: 1rem; background: #dcfce7; border-radius: 0.5rem; border-left: 4px solid #16a534;"><p style="font-weight: 700; margin: 0;">Extensión Horaria</p><p style="font-size: 1.5rem; font-weight: 900; margin: 0; color: #16a534;">9.7%</p></div>
            <div style="padding: 1rem; background: #dcfce7; border-radius: 0.5rem; border-left: 4px solid #16a534;"><p style="font-weight: 700; margin: 0;">Noche</p><p style="font-size: 1.5rem; font-weight: 900; margin: 0; color: #16a534;">4.5%</p></div>
            <div style="padding: 1rem; background: #dcfce7; border-radius: 0.5rem; border-left: 4px solid #16a534;"><p style="font-weight: 700; margin: 0;">Nocturno/Madrugada</p><p style="font-size: 1.5rem; font-weight: 900; margin: 0; color: #16a534;">0.1%</p></div>
        </div>""", unsafe_allow_html=True)
    
    with tab_tasa2:
        col_tab2a, col_tab2b = st.columns(2)
        with col_tab2a:
            st.markdown("""<div style="padding: 1.5rem; background: #fff7ed; border-radius: 0.5rem; text-align: center;"><p style="font-weight: 700; margin: 0;">Urbano</p><p style="font-size: 2rem; font-weight: 900; margin: 0; color: #f59e0b;">15.3%</p></div>""", unsafe_allow_html=True)
        with col_tab2b:
            st.markdown("""<div style="padding: 1.5rem; background: #fff7ed; border-radius: 0.5rem; text-align: center;"><p style="font-weight: 700; margin: 0;">Rural</p><p style="font-size: 2rem; font-weight: 900; margin: 0; color: #f59e0b;">14.6%</p></div>""", unsafe_allow_html=True)
    
    with tab_tasa3:
        st.markdown("""<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 1rem;">
            <div style="padding: 1rem; background: #fff7ed; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">FONASA A</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #f59e0b;">16.2%</p></div>
            <div style="padding: 1rem; background: #eff6ff; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">FONASA D</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #0051d5;">15.3%</p></div>
            <div style="padding: 1rem; background: #eff6ff; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">FONASA C</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #0051d5;">14.8%</p></div>
            <div style="padding: 1rem; background: #dcfce7; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">FONASA B</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #16a534;">14.1%</p></div>
        </div>""", unsafe_allow_html=True)
    
    with tab_tasa4:
        st.markdown("""<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 1rem;">
            <div style="padding: 1rem; background: #fff7ed; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">Adolescencia (12-18)</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #f59e0b;">19.6%</p></div>
            <div style="padding: 1rem; background: #fff7ed; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">Juventud (19-26)</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #f59e0b;">19.5%</p></div>
            <div style="padding: 1rem; background: #fff7ed; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">Infancia (6-11)</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #f59e0b;">19.4%</p></div>
            <div style="padding: 1rem; background: #eff6ff; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">Adultez (27-59)</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #0051d5;">17.6%</p></div>
            <div style="padding: 1rem; background: #dcfce7; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">Primera Infancia (0-5)</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #16a534;">16.7%</p></div>
            <div style="padding: 1rem; background: #dcfce7; border-radius: 0.5rem;"><p style="font-weight: 700; margin: 0;">Tercera Edad (60+)</p><p style="font-size: 1.25rem; font-weight: 900; margin: 0; color: #16a534;">10.5%</p></div>
        </div>""", unsafe_allow_html=True)
    
    st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
    
    st.markdown("""<div style="background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); border-radius: 1rem; padding: 2rem; color: white; margin-bottom: 2rem;"><h4 style="font-size: 1.25rem; font-weight: 700; color: #93c5fd; margin-bottom: 1rem;">💡 Hallazgos Estratégicos Clave</h4><ul style="color: #e5e7eb; line-height: 1.8; font-size: 0.95rem;"><li><strong>Paradoja de Género:</strong> Las mujeres faltan más en volumen (15.8%), pero el modelo predice mejor a los hombres (MASCULINO #2 en importancia)</li><li><strong>El Factor Rural:</strong> PROCEDENCIA_RURAL es el factor #1. Los promedios ocultan que el paciente rural citado a las 08:00 en invierno tiene riesgo crítico</li><li><strong>Extensión Horaria Funciona:</strong> Solo 9.7% de inasistencia vs 16.3% en mañana. Valida la política de horarios extendidos</li></ul></div>""", unsafe_allow_html=True)
    
    st.markdown("""<div style="text-align: center; padding: 3rem 2rem; max-width: 56rem; margin: 0 auto; background: linear-gradient(135deg, #022448 0%, #1e3a5f 100%); border-radius: 1rem; color: white;"><span style="font-size: 3rem; opacity: 0.3;">"</span><p style="font-size: 1.5rem; font-weight: 600; font-style: italic; line-height: 1.5; color: white;">Los datos sin contexto son solo ruido. El modelo detecta los bloques duros que los promedios ocultan.</p><span style="font-size: 3rem; opacity: 0.3;">"</span></div>""", unsafe_allow_html=True)

# ==================== PANTALLA 4: PLANTILLA Y GUÍA ====================
elif st.session_state.current_screen == 'PlantillaGuia':
    st.markdown("""
        <div style="margin-bottom: 2rem;">
            <h1 style="font-size: 2.25rem; font-weight: 900; color: #191c1e;">Plantilla & Guía</h1>
            <p style="color: #43474e; font-size: 1.125rem;">Recursos técnicos para la carga masiva de registros</p>
        </div>
    """, unsafe_allow_html=True)
    
    # INFORMACIÓN GENERAL
    st.markdown("""<div class="info-box" style="margin-bottom: 2rem;">
        <h4 style="font-weight: 700; color: #022448; margin-bottom: 0.5rem;">📋 Acerca de esta Aplicación</h4>
        <p style="color: #43474e; font-size: 0.9rem;">A2S Praedix es un sistema de predicción de inasistencias para el CESFAM Cholchol. Utiliza un modelo Random Forest entrenado con más de 430,000 registros históricos para predecir la probabilidad de que un paciente no asista a su cita agendada.</p>
    </div>""", unsafe_allow_html=True)
    
    # COLUMNAS REQUERIDAS
    st.markdown("""<h3 style="font-size: 1.25rem; font-weight: 700; color: #022448; margin-bottom: 1rem;">📝 Columnas Requeridas</h3>""", unsafe_allow_html=True)
    
    col_info1, col_info2 = st.columns(2)
    with col_info1:
        st.markdown("""<div style="background: white; padding: 1.5rem; border-radius: 1rem; border: 1px solid #e6e8ea; margin-bottom: 1rem;">
        <h5 style="font-weight: 700; color: #022448; margin-bottom: 1rem;">Datos del Paciente</h5>
        <ul style="color: #43474e; font-size: 0.875rem; padding-left: 1.25rem; line-height: 1.8;">
            <li><strong>GENERO:</strong> MASCULINO o FEMENINO</li>
            <li><strong>PROCEDENCIA:</strong> RURAL o URBANO</li>
            <li><strong>ESCOLARIDAD:</strong> BASICA, MEDIA, SUPERIOR, SIN INSTRUCCION, PREBASICA, TECNICO DE NIVEL SUPERIOR</li>
            <li><strong>PREVISION:</strong> FONASA - A, FONASA - B, FONASA - C, FONASA - D</li>
            <li><strong>ES DISCAPACITADA:</strong> SI o NO</li>
            <li><strong>ES SENAME:</strong> SI o NO</li>
            <li><strong>ES EMBARAZADA:</strong> SI o NO</li>
        </ul>
        </div>""", unsafe_allow_html=True)
    
    with col_info2:
        st.markdown("""<div style="background: white; padding: 1.5rem; border-radius: 1rem; border: 1px solid #e6e8ea; margin-bottom: 1rem;">
        <h5 style="font-weight: 700; color: #022448; margin-bottom: 1rem;">Datos de la Hora Agendada</h5>
        <ul style="color: #43474e; font-size: 0.875rem; padding-left: 1.25rem; line-height: 1.8;">
            <li><strong>FECHA ASIGNADA:</strong> Formato fecha (DD/MM/AAAA)</li>
            <li><strong>HORA ASIGNADA:</strong> Formato hora (HH:MM)</li>
            <li><strong>MES:</strong> Enero, Febrero, Marzo, Abril, Mayo, Junio, Julio, Agosto, Septiembre, Octubre, Noviembre, Diciembre</li>
            <li><strong>DIA:</strong> Lunes, Martes, Miércoles, Jueves, Viernes, Sábado, Domingo</li>
            <li><strong>BLOQUE HORARIO:</strong> Mañana, Mediodía, Tarde, Nocturno/Madrugada, Extensión Horaria, Noche</li>
        </ul>
        </div>""", unsafe_allow_html=True)
    
    # PLANTILLA DESCARGABLE CON DROPDOWNS
    st.markdown("""<h3 style="font-size: 1.25rem; font-weight: 700; color: #022448; margin-bottom: 1rem;">⬇️ Descargar Plantilla</h3>""", unsafe_allow_html=True)
    
    wb_plantilla = crear_plantilla_excel()
    buffer_plantilla = io.BytesIO()
    wb_plantilla.save(buffer_plantilla)
    buffer_plantilla.seek(0)
    plantilla_data = buffer_plantilla.getvalue()
    
    st.markdown("""<p style="color: #43474e; margin-bottom: 1rem;">Descargue la plantilla con el formato correcto para cargar sus archivos. La plantilla incluye datos de ejemplo que puede reemplazar.</p>""", unsafe_allow_html=True)
    
    st.download_button(
        label="📥 Descargar Plantilla Excel",
        data=plantilla_data,
        file_name="plantilla_some_cholchol.xlsx",
        type="primary"
    )
    
    # PREGUNTAS FRECUENTES
    st.markdown("""<h3 style="font-size: 1.25rem; font-weight: 700; color: #022448; margin-bottom: 1rem; margin-top: 2rem;">❓ Preguntas Frecuentes</h3>""", unsafe_allow_html=True)
    
    with st.expander("¿Qué formato de archivo es compatible?"):
        st.markdown("""<p style="color: #43474e;">El sistema acepta archivos en formato <strong>.xlsx</strong> (Excel) y <strong>.csv</strong>. El tamaño máximo recomendado es de 50MB por archivo.</p>""", unsafe_allow_html=True)
    
    with st.expander("¿Cómo obtengo la plantilla?"):
        st.markdown("""<p style="color: #43474e;">Use el botón de descarga "Descargar Plantilla Excel" en esta sección para obtener la plantilla con el formato correcto.</p>""", unsafe_allow_html=True)
    
    with st.expander("¿Qué datos necesito para cada paciente?"):
        st.markdown("""<p style="color: #43474e;">Para cada paciente debe proporcionar: género, procedencia (rural/urbano), escolaridad, previsión de salud, fecha y hora de la cita agendada. Los campos de condiciones especiales son opcionales.</p>""", unsafe_allow_html=True)
    
    with st.expander("¿Qué significa cada nivel de riesgo?"):
        st.markdown("""<p style="color: #43474e;"><strong>🔴 Riesgo Crítico (≥80%):</strong> Contactar inmediatamente por teléfono.<br>
        <strong>🟡 Riesgo Alto (60-79%):</strong> Enviar recordatorio por WhatsApp.<br>
        <strong>🟢 Riesgo Bajo (&lt;60%):</strong> No es necesario contactar.</p>""", unsafe_allow_html=True)
    
    with st.expander("¿Cómo se protege la privacidad de los pacientes?"):
        st.markdown("""<p style="color: #43474e;">Los datos se procesan únicamente de forma local en su navegador. No se almacenan en servidores externos. El modelo fue entrenado con datos históricos anonimizados.</p>""", unsafe_allow_html=True)
    
    with st.expander("¿Puedo usar el sistema sin conexión a internet?"):
        st.markdown("""<p style="color: #43474e;">Una vez cargada la página, el sistema funciona completamente sin conexión. Sin embargo, necesita internet para acceder inicialmente a la aplicación.</p>""", unsafe_allow_html=True)

# Footer con logo centrado
st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)
st.markdown("---")

_, footer_center, _ = st.columns([1, 2, 1])
with footer_center:
    st.markdown("""
    <div style="text-align: center;">
        <p style="font-size: 1rem; font-weight: 600; color: #022448; margin-bottom: 0.5rem;">Aplicación desarrollada por Alain Antinao Sepúlveda</p>
        <p style="font-size: 0.875rem; color: #43474e; margin-bottom: 0.5rem;">📧 Contacto: alain.antinao.s@gmail.com</p>
        <p style="font-size: 0.875rem; color: #0051d5;">🌐 <a href="https://alain-antinao-s.notion.site/Alain-C-sar-Antinao-Sep-lveda-1d20a081d9a980ca9d43e283a278053e" style="color: #0051d5; text-decoration: none;">Más información</a></p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<div style='height: 1rem;'></div>", unsafe_allow_html=True)